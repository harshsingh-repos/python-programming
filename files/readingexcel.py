import openpyxl

path = "/Users/harsh/Documents/GitHub-Harsh/python-programming/files/demo_data.xlsx"
def openExcel(file):
    df = openpyxl.load_workbook(file)
    return df


def readExcel(file):
    df = openExcel(file)
    dataRead = df.active
    for row in range(0, dataRead.max_row):
        for col in dataRead.iter_cols(1, dataRead.max_column):
            print(col[row].value)


readExcel(path)